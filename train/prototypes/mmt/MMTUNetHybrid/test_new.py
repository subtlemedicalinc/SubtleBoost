from networks.mmt import MMT as generator
from configs.config import get_config
from evaluator import split_data
import argparse
import os
import glob
import numpy as np
import torch
import pickle
import lpips
from skimage.metrics import structural_similarity as ssim
from skimage.metrics import peak_signal_noise_ratio as psnr
from skimage.metrics import mean_squared_error as mse
from sklearn.metrics import mean_absolute_error as mae
from torchvision.utils import save_image
from tqdm import tqdm
from configs.config import get_config
from utils import make_image_grid
from openpyxl import Workbook

loss_fn_vgg = lpips.LPIPS(net='vgg').cuda()

input_combination_brats = [[0], [1], [2], [3], [0, 1], [0, 2], [0, 3], [1, 2], [1, 3], [2, 3], [0, 1, 2],
                           [0, 1, 3], [0, 2, 3], [1, 2, 3]]

input_combination_3 = [[1, 2, 3], [0, 2, 3], [0, 1, 3], [0, 1, 2]]

input_combination_zerogad = [[0, 2, 3]]

input_combination_ixi = [[0], [1], [2], [0, 1], [0, 2], [1, 2]]

input_combination_2 = [[1, 2], [0, 2], [0, 1]]


def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--data_path', type=str,
                        default='/mnt/raid/jiang/projects/SubtleGAN/data/brats2021_slices_crop192x160',
                        help='root dir for data')
    parser.add_argument('--cfg', type=str, default='configs/mmt.yml')
    parser.add_argument('--dataset', type=str,
                        default='BRATS', help='experiment_name')
    parser.add_argument('--n_gpu', type=int, default=1, help='total gpu')
    parser.add_argument('--k', type=int,
                        default=None, help='number of inputs')
    parser.add_argument('--zero_gad', action='store_true', help='eval zero_gad')
    parser.add_argument('--ckpt', type=str, default=None)
    parser.add_argument('--model_path', type=str, default=None)
    parser.add_argument('--vis', action='store_true', help='visualize results')
    parser.add_argument('--seg', action='store_true', help='test seg acc')
    parser.add_argument('--masked', action='store_true', help='test similarity within tumor mask')
    parser.add_argument('--save_dir', type=str, default='evaluation')
    parser.add_argument('--batch_size', type=int, default=32)
    parser.add_argument('--seed', type=int,
                        default=1234, help='random seed')
    parser.add_argument('--n_contrast', type=int, default=4, help='total number of contrast in the dataset')

    return parser.parse_args()


def list2str(x):
    f = ''
    for i in x:
        f += f'{i}'
    return f


def input2str(x, n):
    f = ''
    for i in range(n):
        if i in x:
            f += f'{1}'
        else:
            f += f'{0}'
    return f


def get_output(x, n):
    if n == 3:
        contrast_dict = ['T1', 'T2', 'PD']
    else:
        contrast_dict = ['T1', 'T1Gd', 'T2', 'FLAIR']

    f = [contrast_dict[x_i] for x_i in x]
    return f


def lpips_metrics(img1, img2):
    img1 = torch.from_numpy(img1).unsqueeze(0).unsqueeze(0).cuda()
    img2 = torch.from_numpy(img2).unsqueeze(0).unsqueeze(0).cuda()
    img1 = torch.cat((img1, img1, img1), dim=1)
    img2 = torch.cat((img2, img2, img2), dim=1)
    img1 = 2 * img1 - 1
    img2 = 2 * img2 - 1
    score = loss_fn_vgg(img1, img2)
    return score.item()


class AverageMeter(object):
    """Computes and stores the average and current value"""

    def __init__(self):
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update(self, val, n=1):
        self.val = val
        self.sum += val * n
        self.count += n
        self.avg = self.sum / self.count


def evaluator(model, inputs, targets, cases, save_dir=None, n_contrast=3, vis=False):
    input_tag = input2str(inputs, n_contrast)
    model.eval()

    output_contrasts = get_output(targets, n_contrast)
    wb = Workbook()
    ws = [wb.create_sheet(title=output_contrast) for output_contrast in output_contrasts]
    for sheet in ws:
        sheet.cell(1, 1).value = 'Filename'
        sheet.cell(1, 2).value = 'MSE'
        sheet.cell(1, 3).value = 'MAE'
        sheet.cell(1, 4).value = 'SSIM'
        sheet.cell(1, 5).value = 'PSNR'
        sheet.cell(1, 6).value = 'LPIPS'

    metrics_meters = [[AverageMeter() for _ in range(5)] for _ in range(len(targets))]

    row = 2
    with torch.no_grad():
        for case in tqdm(cases):
            case_name = case.split("/")[-1]
            if vis:
                vis_dir = os.path.join(save_dir, input_tag, case_name)
                os.makedirs(vis_dir, exist_ok=True)

            slices = sorted(glob.glob(os.path.join(case, "*.npy")))

            for file in slices:
                data = np.load(file)
                n_channel = data.shape[0]
                image = torch.from_numpy(data.astype(np.float32)).unsqueeze(0)
                data = [image[:, i, :, :].unsqueeze(0) for i in range(n_channel)]  # [(1, 1, H, W)]
                img_inputs, img_targets, contrast_input, contrast_output = split_data(data, inputs, targets)
                img_outputs, _, _ = model(img_inputs, contrast_input, contrast_output)

                # save visualization results
                if vis:
                    slice_num = file.split("/")[-1].split(".")[0]
                    save_image(make_image_grid(img_inputs), f'{vis_dir}/{slice_num}_input.png')
                    save_image(make_image_grid(img_outputs), f'{vis_dir}/{slice_num}_output.png')
                    save_image(make_image_grid(img_targets), f'{vis_dir}/{slice_num}_gt.png')

                # for each contrast
                for i, outputs in enumerate(img_outputs):
                    output_imgs = outputs.detach().cpu().numpy()
                    target_imgs = img_targets[i].detach().cpu().numpy()

                    img_o = output_imgs[0, 0, :, :]
                    img_t = target_imgs[0, 0, :, :]
                    img_max = img_t.max()
                    img_o /= img_max
                    img_t /= img_max
                    img_o = img_o.clip(0, 1)

                    mse_score = mse(img_o, img_t)
                    mae_score = mae(img_o, img_t)
                    ssim_score = ssim(img_o, img_t)
                    psnr_score = psnr(img_t, img_o)
                    lpips_score = lpips_metrics(img_t, img_o)

                    metrics_meters[i][0].update(mse_score)
                    metrics_meters[i][1].update(mae_score)
                    metrics_meters[i][2].update(ssim_score)
                    metrics_meters[i][3].update(psnr_score)
                    metrics_meters[i][4].update(lpips_score)

                    slice_name = file.split("/")
                    slice_name = f'{slice_name[-2]}/{slice_name[-1]}'
                    ws[i].cell(row, 1).value = slice_name
                    ws[i].cell(row, 2).value = mse_score
                    ws[i].cell(row, 3).value = mae_score
                    ws[i].cell(row, 4).value = ssim_score
                    ws[i].cell(row, 5).value = psnr_score
                    ws[i].cell(row, 6).value = lpips_score

                row += 1

    metrics = [{} for _ in range(len(targets))]
    for i in range(len(targets)):
        metrics[i]['mse'] = metrics_meters[i][0].avg
        metrics[i]['mae'] = metrics_meters[i][1].avg
        metrics[i]['ssim'] = metrics_meters[i][2].avg
        metrics[i]['psnr'] = metrics_meters[i][3].avg
        metrics[i]['lpips'] = metrics_meters[i][4].avg

    print(f"***Inputs: {inputs}; Outputs: {targets}; {metrics}")

    wb_fn = os.path.join(save_dir, f'{input_tag}.xlsx')
    print(wb_fn)
    wb.save(wb_fn)
    return metrics


def evaluate(args, model, input_combination, split='test'):
    data_dir = os.path.join(args.data_path, split)
    if args.dataset == 'BRATS':
        cases = sorted(glob.glob(f"{data_dir}/Bra*"))
    else:
        cases = sorted(glob.glob(f"{data_dir}/IXI*"))

    save_dir = os.path.join(args.model_path, args.save_dir)
    os.makedirs(save_dir, exist_ok=True)

    metrics = []
    for inputs in input_combination:
        targets = list(set(range(args.n_contrast)) - set(inputs))
        print(f"***Inputs: {inputs}; Outputs: {targets}")
        metric = evaluator(model, inputs, targets, cases, save_dir, args.n_contrast, args.vis)
        metrics.append(metric)
    return metrics


if __name__ == "__main__":
    args = get_args()
    args.n_contrast = 4 if args.dataset == 'BRATS' else 3
    args.cfg = 'configs/mmt.yml' if args.dataset == 'BRATS' else 'configs/mmt_ixi.yml'
    
    config = get_config(args)
    G = generator(img_size=config.DATA.IMG_SIZE,
                  patch_size=config.MODEL.SWIN.PATCH_SIZE,
                  in_chans=config.MODEL.SWIN.IN_CHANS,
                  out_chans=config.MODEL.SWIN.OUT_CHANS,
                  embed_dim=config.MODEL.SWIN.EMBED_DIM,
                  depths=config.MODEL.SWIN.DEPTHS,
                  num_heads=config.MODEL.SWIN.NUM_HEADS,
                  window_size=config.MODEL.SWIN.WINDOW_SIZE,
                  mlp_ratio=config.MODEL.SWIN.MLP_RATIO,
                  qkv_bias=config.MODEL.SWIN.QKV_BIAS,
                  qk_scale=config.MODEL.SWIN.QK_SCALE,
                  drop_rate=config.MODEL.DROP_RATE,
                  drop_path_rate=config.MODEL.DROP_PATH_RATE,
                  ape=config.MODEL.SWIN.APE,
                  patch_norm=config.MODEL.SWIN.PATCH_NORM,
                  use_checkpoint=config.TRAIN.USE_CHECKPOINT,
                  seg=args.seg,
                  num_contrast=args.n_contrast).cuda()
    state_dict = torch.load(os.path.join(args.model_path, args.ckpt), map_location='cpu')
    G.load_state_dict(state_dict['G'])
    G.eval()
    for param in G.parameters():
        param.requires_grad = False

    if args.zero_gad:
        input_combination = input_combination_zerogad
    elif args.k == 3:
        input_combination = input_combination_3
    elif args.k == 2:
        input_combination = input_combination_2
    else:
        input_combination = input_combination_brats if args.dataset == 'BRATS' else input_combination_ixi

    metrics = evaluate(args, G, input_combination, split='test')

    filename = os.path.join(args.model_path, args.save_dir, "metrics.pkl")
    pickle.dump({'metrics': metrics, 'inputs': input_combination}, open(filename, 'wb'))

    with open(os.path.join(args.model_path, args.save_dir, "test_results.txt"), "w") as f:
        for i in range(len(input_combination)):
            output_combination = list(set(range(args.n_contrast)) - set(input_combination[i]))
            for j in range(len(metrics[i])):
                for m in ['ssim', 'mae', 'psnr', 'mse', 'lpips']:
                    msg = f'test_{list2str(input_combination[i])}/{m}_{output_combination[j]}: {metrics[i][j][m]}\n'
                    f.write(msg)
                    print(msg)
