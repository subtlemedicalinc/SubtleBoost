from modules.advanced_gans.models import *
from modules.helpers import create_dataloaders, Resize, ToTensor
from skimage.metrics import structural_similarity as ssim
from skimage.metrics import peak_signal_noise_ratio as psnr
from skimage.metrics import mean_squared_error as mse
from sklearn.metrics import mean_absolute_error as mae
from torchvision.utils import make_grid, save_image
from kornia.geometry.transform import resize
from tqdm import tqdm
from openpyxl import Workbook
import itertools
import torch
import os
import torch
import argparse
import pdb
import cv2
import glob
import lpips
import pickle


loss_fn_vgg = lpips.LPIPS(net='vgg')


def lpips_metrics(img1, img2):
    img1 = torch.from_numpy(img1).unsqueeze(0).unsqueeze(0)
    img2 = torch.from_numpy(img2).unsqueeze(0).unsqueeze(0)
    img1 = torch.cat((img1, img1, img1), dim=1)
    img2 = torch.cat((img2, img2, img2), dim=1)
    img1 = 2*img1 - 1
    img2 = 2*img2 - 1
    score = loss_fn_vgg(img1, img2)
    return score.item()
    

class AverageMeter(object):
    """Computes and stores the average and current value"""
    def __init__(self):
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update(self, val, n=1):
        self.val = val
        self.sum += val * n
        self.count += n
        self.avg = self.sum / self.count

def make_image_grid(image_array):
    images = []
    n = image_array.shape[0]
    for i in range(n):
        image = cv2.resize(image_array[i], (192, 160))
        image = torch.from_numpy(image).unsqueeze(0)
        image = (image - image.min()) / (image.max() - image.min())
        image = image.repeat(3,1,1)
        image = torch.rot90(image, 3, [1, 2])
        images.append(image)
    image_grid = make_grid(images, padding=0)
    return image_grid

def list2str(x):
    f = ''
    for x_i in x:
        f += f'{x_i}'
    return f

# def list2str(x):
#     f = ''
#     for i, x_i in enumerate(x):
#         if x_i == 1:
#             f += f'{i}'
#     return f

def get_output(x):
    contrast_dict = ['T1', 'T1Gd', 'T2', 'FLAIR']
    f = []
    for i, x_i in enumerate(x):
        if x_i == 0:
            f.append(f'{contrast_dict[i]}')
    return f


def eval_brats(generator, scenario, patient_list, model_path, vis_dir=None, vis=False):
    scenario_tag = list2str(scenario)
    scenario = np.array(scenario)
    generator.eval()
    impute_tensor = torch.zeros((1, 256, 256), device='cuda')
    metrics_meters = [[AverageMeter() for _ in range(5)] for _ in range(np.sum(scenario==0))]
    
    output_contrasts = get_output(scenario)
    wb = Workbook()
    ws = [wb.create_sheet(title=output_contrast) for output_contrast in output_contrasts]
    for sheet in ws:
        sheet.cell(1, 1).value = 'Filename'
        sheet.cell(1, 2).value = 'MSE'
        sheet.cell(1, 3).value = 'MAE'
        sheet.cell(1, 4).value = 'SSIM'
        sheet.cell(1, 5).value = 'PSNR'
        sheet.cell(1, 6).value = 'LPIPS'
    
    # pdb.set_trace()
    row = 2
    with torch.no_grad():
        for patient in tqdm(patient_list):
            pat_name = patient.split("/")[-1]
            #print(f'Scenario {scenario_tag}: evaluating {pat_name}')
            if vis:
                save_dir = os.path.join(model_path, vis_dir, scenario_tag, pat_name)
                os.makedirs(save_dir, exist_ok=True)
            
            patient_slices = sorted(glob.glob(os.path.join(patient, "*.npy")))
            for patient_slice in patient_slices:
                
                slice_name = patient_slice.split("/")
                slice_name = f'{slice_name[-2]}/{slice_name[-1]}'
                
                i = patient_slice.split("/")[-1].split(".")[0]
                data = np.load(patient_slice)
                
                x_test_r = torch.from_numpy(data).unsqueeze(0).cuda().float()
                targets = x_test_r.detach().cpu().numpy()[0, scenario==0, :] 
                x_test_r = resize(x_test_r, (256, 256), align_corners=False).detach()
                x_test_z = x_test_r.clone().cuda().type(torch.cuda.FloatTensor)
                for idx_, k in enumerate(scenario):
                    if k == 0:
                        x_test_z[:, idx_, ...] = impute_tensor
                G_result = generator(x_test_z)
                outputs = G_result.detach().cpu().numpy()[0, scenario==0, :]
                
                inputs = x_test_r.detach().cpu().numpy()[0, scenario==1, :] 
                
                # save images
                if vis:
                    save_image(make_image_grid(inputs), f'{save_dir}/{i}_input.png')
                    save_image(make_image_grid(outputs), f'{save_dir}/{i}_output.png')
                    save_image(make_image_grid(targets), f'{save_dir}/{i}_gt.png')
                
                #targets = data[scenario==0, :].copy()
                #pdb.set_trace()
                for j in range(outputs.shape[0]):
                    img_o = outputs[j, :, :]
                    img_t = targets[j, :, :]
                    
                    img_o = cv2.resize(img_o, (192, 160))
                    
                    #img_o = img_o.astype(np.float32)
                    #img_t = img_t.astype(np.double)
                    #img_t = cv2.resize(img_t, (192, 160))
                    
                    img_max = img_t.max()
                    img_o /= img_max
                    img_t /= img_max
                    img_o = img_o.clip(0, 1)
                    
                    #pdb.set_trace()
                    
                    mse_score = mse(img_o, img_t)
                    mae_score = mae(img_o, img_t)
                    ssim_score = ssim(img_o, img_t)
                    psnr_score = psnr(img_t, img_o)
                    lpips_score = lpips_metrics(img_t, img_o)
                    
                    metrics_meters[j][0].update(mse_score)
                    metrics_meters[j][1].update(mae_score)
                    metrics_meters[j][2].update(ssim_score)
                    metrics_meters[j][3].update(psnr_score)
                    metrics_meters[j][4].update(lpips_score)
                    
                    ws[j].cell(row, 1).value = slice_name
                    ws[j].cell(row, 2).value = mse_score
                    ws[j].cell(row, 3).value = mae_score
                    ws[j].cell(row, 4).value = ssim_score 
                    ws[j].cell(row, 5).value = psnr_score
                    ws[j].cell(row, 6).value = lpips_score
                    
                row += 1
    metrics = [{} for _ in range(np.sum(scenario==0))]
    for i in range(np.sum(scenario==0)):
        metrics[i]['mse'] = metrics_meters[i][0].avg
        metrics[i]['mae'] = metrics_meters[i][1].avg
        metrics[i]['ssim'] = metrics_meters[i][2].avg
        metrics[i]['psnr'] = metrics_meters[i][3].avg
        metrics[i]['lpips'] = metrics_meters[i][4].avg
    print(f"***{scenario} {metrics}")
    wb_fn = os.path.join(model_path, vis_dir, f'{scenario_tag}.xlsx')
    print(wb_fn)
    wb.save(wb_fn)
    return metrics
                

def evaluator_brats(model, scenarios, patient_list, model_path, vis_dir='vis', vis=False):
    metrics = []
    for scenario in scenarios:
        print(f"***Scenario: {scenario}")
        metric = eval_brats(model, scenario, patient_list, model_path, vis_dir=vis_dir, vis=vis)
        metrics.append(metric)
    return metrics



parser = argparse.ArgumentParser()
parser.add_argument('--gpu', type=str, default='0')
parser.add_argument('--n_contrast', type=int, default=3)
parser.add_argument('--n_slices', type=int, default=90)
parser.add_argument('--n_cases', type=int, default=30)
parser.add_argument('--single', action='store_true', help='single missing data')
parser.add_argument('--ckpt', type=str, default="model/MMGAN/mmgan_ixi_single/generator_param_mmgan_ixi_single_60.pkl")
parser.add_argument('--data_path', type=str, default="/mnt/raid/jiang/projects/SubtleGAN/data/IXI/hdf5/")
parser.add_argument('--model_path', type=str, default="model/MMGAN/mmgan_ixi_single/")
parser.add_argument('--save_dir', type=str, default="vis")
parser.add_argument('--vis', action='store_true', default=False)
args = parser.parse_args()

os.environ['CUDA_VISIBLE_DEVICES'] = args.gpu
generator = GeneratorUNet(in_channels=args.n_contrast, out_channels=args.n_contrast, with_relu=True, with_tanh=False)
ckpt = torch.load(args.ckpt, map_location='cpu')
generator = nn.DataParallel(generator.cuda())
generator.load_state_dict(ckpt['state_dict'])


test_patient = sorted(glob.glob(os.path.join(args.data_path, "BraTS*")))

if args.single:
    scenarios = [[1, 1, 1, 0], [1, 1, 0, 1], [1, 0, 1, 1], [0, 1, 1, 1]]
else:
    scenarios = list(map(list, itertools.product([0, 1], repeat=4)))
    scenarios.remove([0,0,0,0])
    scenarios.remove([1,1,1,1])
    
metrics = evaluator_brats(generator, scenarios, test_patient, args.model_path, vis_dir=args.save_dir, vis=args.vis)

filename = os.path.join(args.model_path, args.save_dir, "metrics.pkl")
input_combination = []
for scenario in scenarios:
    inputs = []
    for i, num in enumerate(scenario):
        if num == 1:
            inputs.append(i)
    input_combination.append(inputs)
pickle.dump({'metrics': metrics, 'inputs': input_combination, 'scenarios': scenarios}, open(filename, 'wb'))

with open(os.path.join(args.model_path, args.save_dir, "test_results.txt"), "w") as f:
     for i, scenario in enumerate(scenarios):
        scenario_tag = list2str(scenario)
        for j in range(len(metrics[i])):
            for m in ['ssim', 'mae', 'psnr', 'mse', 'lpips']:
                msg = f'test_{scenario_tag}/{m}_{j}: {metrics[i][j][m]}\n'
                f.write(msg)
                print(msg)
